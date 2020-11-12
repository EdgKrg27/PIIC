# script para realizar el procesamiento de imgaen al igual que toda la estadística

# librerias
import numpy as np
import cv2 as cv
import pandas as pd
import scipy.stats as ss
from openpyxl import load_workbook
from math import pi
from matplotlib import pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from matplotlib.figure import Figure
import tkinter as tk
from tkinter import messagebox as mb


# FUNCIÓN QUE CIERRA TODAS LAS VENTANAS DE LAS IMAGENES
def func_cerrar():
    cv.destroyAllWindows()


class procesamiento:
    # VARIABLES GLOBALES
    aux1, aux2, aux3, aux4 = None, None, None, None
    escala = None
    res = None
    copia_imagen_original = None

    def nada(self, val):
        pass

    # FUNCIÓN QUE BUSCA, ABRE Y RECORTA LA IMAGEN
    def abrir_imagen(self, img):
        imagen_original = cv.imread(img)  # LECTURA DE IMAGEN
        self.copia_imagen_original = imagen_original.copy()  # COPIA DE LA IMAGEN ORIGINAL
        gris = cv.cvtColor(self.copia_imagen_original, cv.COLOR_BGR2GRAY)  # IMAGEN EN ESCALA DE GRISES

        # RECORTE DE IMAGEN PARA MEJOR PROCESAMIENTO
        gris_copia = gris.copy()  # COPIA DE LA IMAGENE GRIS PARA RECORTAR
        alto, ancho = gris_copia.shape[0:2]  # TAMAÑO DE LA IMAGEN EN PIXELES
        inicio_renglon = int(alto * 0)  # OPERACIONES PARA RECORTAR LA IMAGEN
        inicio_columna = int(ancho * 0)
        fin_renglon = int(alto * .94)
        fin_columna = int(ancho)
        imagen_nueva = gris_copia[inicio_renglon:fin_renglon, inicio_columna:fin_columna]  # MAPEO DEL NUEVO TAMAÑO

        cv.imshow('Imagen', gris)  # IMPRIME LA IMAGEN EN PANTALLA
        self.aux1 = imagen_nueva

    #FUNCIÓN QUE CAMBIAR EL CONTRASTE DE LA IMAGEN
    def contraste(self, val):
        imagen = self.aux1
        cv.destroyAllWindows()  # CIERRA LA IMAGEN ANTERIOR
        img_contraste = cv.addWeighted(imagen.copy(), val / 3, np.zeros(imagen.shape, imagen.dtype), 0, 0)  #CAMBIA EL CONTRASTE
        cv.imshow('Contraste', img_contraste)
        self.aux2 = img_contraste

    # FUNCIÓN QUE DESENFOCA LA IMAGEN
    def desenfoque(self, val):
        imagen = self.aux2
        cv.destroyAllWindows()
        # imagen_borrosa = cv.bilateralFilter(imagen, 9, val, val)
        imagen_borrosa = cv.medianBlur(imagen, val) # DESENFOCA LA IMAGEN
        cv.imshow('Desenfoque', imagen_borrosa)
        self.aux3 = imagen_borrosa

    # FUNCIÓN CAMBIA A LA IMAGEN EN SU FORMA BINARIA
    def imgBinary(self):
        imagen = self.aux3
        cv.destroyAllWindows()
        _, img_binario = cv.threshold(imagen, 150, 255, cv.THRESH_BINARY)  # CAMBIA LA IMAGEN CON FONDO NEGRO CIRCULOS BLANCOS
        cv.imshow('Binario', img_binario)
        self.aux4 = img_binario

    # FUNCIÓN QUE RECIBE LA OPCIÓN DE LA ESCALA
    def opcion_escala(self, scale):
        self.escala = scale

    # FUNCIÓN QUE DETECTA LOS CÍRCULOS Y CALCULA LAS ÁREAS DE LOS MISMOS
    def deteccion(self):
        try:
            imagen = self.aux4
            cv.destroyAllWindows()
            _, contours, _ = cv.findContours(imagen, cv.RETR_EXTERNAL, cv.CHAIN_APPROX_SIMPLE)  # ENCUENTRA LOS CONTORNOS
            circulos = [cv.minEnclosingCircle(cnt) for cnt in contours]  # IDENTIFICACIÓN DE CÍRCULOS
            #print(f"Circulos encontrados: {len(circulos)}, Círculos: {circulos}")
            areas_gen = (pi * (radio * radio) * self.escala for _, radio in circulos)  #  CALCULA LAS ÁREAS DE LOS CÍRCULOS ENCONTRADOS
            #self.res = pd.DataFrame({'Areas': areas_gen})
            self.res = pd.Series(areas_gen)  # CREA UN DATA FRAME DE LAS ÁREAS
            #print(self.res)

            # DIBUJA LOS CIRCULOS ENCONTRADOS EN LA IMAGEN ORIGINAL
            i = 0
            font = cv.FONT_HERSHEY_SIMPLEX
            for c in contours:
                M = cv.moments(c)
                if M['m00'] == 0:
                    M['m00'] = 1
                x = int(M['m10'] / M['m00'])
                y = int(M['m01'] / M['m00'])
                message = str(i + 1)
                cv.putText(self.copia_imagen_original, message, (x - 40, y), font, 0.5, (0, 0, 255), 1)
                cv.drawContours(self.copia_imagen_original, [c], -1, (0, 0, 255), 1)
                i = i + 1

            cv.imshow('Detección', self.copia_imagen_original)
            return self.res
        except(cv.error):
            mb.showerror('Error', "Dato incorrecto")

    # FUNCIONES ESTADÍSTICAS
    def areas(self):
        data = np.round(self.res.to_numpy(), 4)
        return data

    def mediana(self):
        mu = np.round(self.res.mean(), 4)
        #print("Mediana: {}".format(mu))
        return mu

    def desviacion_estandar(self):
        sigma = np.round(self.res.std(), 4)
        #print("Desviación Estandar: {}".format(sigma))
        return sigma

    def minimo(self):
        mi = np.round(self.res.min(), 4)
        #print("Minimo: {}".format(mi))
        return mi

    def maximo(self):
        ma = np.round(self.res.max(), 4)
        #print("Maximo: {}".format(ma))
        return ma

    def areas_ordenadas(self):
        ord = np.round(self.res.to_numpy(), 4)
        ord.sort()
        #print(ord[::-1])
        return ord[::-1]

    def conteo(self, tam):
        data = self.res.to_numpy()
        mi = self.minimo()
        ma = self.maximo()
        semilla = np.round(np.linspace(mi, ma, tam+1), 4)
        contador, _ = np.histogram(data, semilla)
        self.bin = semilla
        self.cnt = contador
        """print("\nIntervalo\tResultado")
        for i, j, count in zip(semilla, np.roll(semilla, -1), contador):
            print(f"{f'{i}-{j}'}\t{count}")"""
        return semilla, contador

    # GRÁFICAS SE MUESTRAN EN PANTALLA
    def grafica_conteo(self):
        try:
            data = self.res.to_numpy()
            semilla = self.bin
            plt.hist(data, semilla, rwidth=0.5, color='red')
            plt.title('Conteo de Particulas')
            plt.xlabel('Tamaño')
            plt.ylabel('Cantidad')
            plt.savefig("docs/conteo.png", bbox_inches='tight')  # GUARDA LA GRAFICA COMO IMAGEN TIPO .PNG
            winGC = tk.Tk()
            winGC.wm_title("Gráfica de conteo de particulas")
            fig = Figure(figsize=(5, 4), dpi=100)
            canvas = FigureCanvasTkAgg(fig, master=winGC)
            canvas.get_tk_widget().grid(row=0, column=0, rowspan=6)
            p = fig.gca()
            p.hist(data, semilla, rwidth=0.5, color='red')  # CREACIÓN DEL HISTOGRAMA DEL CONTEO DE LOS DATOS
            p.set_xlabel('Tamaño')
            p.set_ylabel('Cantidad')
            p.set_title("Conteo de particulas")
            canvas.draw()
            bot = tk.Button(winGC, text='Cerrar', command=winGC.destroy)
            bot.grid(row=7, column=0)
        except(AttributeError):
            mb.showerror('Error', "No se puede dibujar la gráfica, datos incorrectos")

    def grafica_normal(self):
        try:
            semilla = self.bin
            mu = self.mediana()
            sigma = self.desviacion_estandar()
            normal = ss.norm(mu, sigma)  # CALCULO DE LA NORMAL
            fp = normal.pdf(semilla)  # REALIZANDO UNA FUNCIÓN DE PROBABILIDAD
            plt.plot(semilla, fp, color='red')
            plt.title('Distribución Normal')
            plt.xlabel('Tamaño')
            plt.ylabel('Distribución')
            plt.savefig("docs/normal.png", bbox_inches='tight')  # GURDAMOS LA GRAFICA COMO IMAGEN TIPO .PNG
            winGN = tk.Tk()
            winGN.wm_title("Gráfica de la distribución normal")
            fig = Figure(figsize=(5, 4), dpi=100)
            # fig.add_subplot(111).plot(semilla, fp, color='red')
            canvas = FigureCanvasTkAgg(fig, master=winGN)
            canvas.get_tk_widget().grid(row=0, column=0, rowspan=6)
            p = fig.gca()
            p.plot(semilla, fp, color='red')  # CREACIÓN DE LA GRAFICA
            p.set_xlabel('Tamaño')
            p.set_ylabel('Distribución')
            p.set_title("Distribución Normal")
            canvas.draw()
            bot = tk.Button(winGN, text='Cerrar', command=winGN.destroy)
            bot.grid(row=7, column=0)
        except(AttributeError):
            mb.showerror('Error', "No se puede dibujar la gráfica, datos incorrectos")

    # FUNCIÓN QUE CREA EL EXCEL FINAL
    def crear_excel(self):
        file = 'docs/medidas.xlsx'  # NOMBRE DEL EXCEL A CREAR
        semilla = self.bin
        contador = self.cnt

        # DATOS ESTADÍSTICOS
        data_estd = [self.mediana(), self.desviacion_estandar(), self.minimo(), self.maximo()]
        indices = ['Media', 'Desviación Estandar', 'Mínimo', 'Máximo']

        data = self.res #Áreas

        ord = pd.DataFrame({'Ordenamiento': self.areas_ordenadas()}) # Áreas ordenadas de menor a mayor

        # Datos de la normal
        r1 = pd.DataFrame(data=indices, columns=[' '])
        r2 = pd.DataFrame(data=data_estd, columns=['Valores'])
        b1 = pd.DataFrame({'Inferior': semilla})
        b2 = pd.DataFrame({'Superior': np.roll(semilla, -1)})
        cnt = pd.DataFrame({'Conteo': contador})

        # CREACIÓN DEL ARCHIVO DE EXCEL
        writer = pd.ExcelWriter(file, engine='openpyxl')
        data.to_excel(writer)
        writer.save()

        wb = load_workbook(writer)
        sheet = wb.active
        ws = wb['Sheet1']
        sheet['A1'] = 'Áreas'

        for index, row in r1.iterrows():
            cell = 'C%d' % (index + 2)
            ws[cell] = row[0]

        sheet['D1'] = 'Valores'
        for index, row in r2.iterrows():
            cell = 'D%d' % (index + 2)
            ws[cell] = row[0]

        sheet['E1'] = 'Ordenamiento'
        for index, row in ord.iterrows():
            cell = 'E%d' % (index + 2)
            ws[cell] = row[0]

        sheet['G1'] = 'Intervalo'
        for index, row in b1.iterrows():
            cell = 'G%d' % (index + 2)
            ws[cell] = row[0]

        for index, row in b2.iterrows():
            cell = 'H%d' % (index + 2)
            ws[cell] = row[0]

        sheet['I1'] = 'Conteo'
        for index, row in cnt.iterrows():
            cell = 'I%d' % (index + 2)
            ws[cell] = row[0]

        wb.save(file)  # GUARDA EL ARCHIVO EXCEL